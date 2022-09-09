# -*- coding: utf-8 -*-
"""
Created on Fri Aug 26 08:30:00 2022

@author: Jhonatan Martínez
"""

import io
from flask import Response, redirect, url_for
from openpyxl import Workbook


class Excel:
    """
    Esta clase permite realizar la descarga de un archivo de Excel
    """

    @staticmethod
    def download(columns, rows, file_name):
        """
        Permite realizar la descarga de un archivo de Excel desde una consulta a una base de datos

        Params:
            * **columns (List):** Datos con los headers o nombres de las columnas. \n
            * **rows (List):** Datos a colocar en el archivo. \n
            * **file_name (Str):** Nombre con el que se va a descargar el archivo. \n

        Returns:

        """
        # Variable output para guardar el archivo en formato binario
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        row_number = 1
        column_number = 1
        # Con este for se llenan los headers o encabezados del archivo
        for column in columns:
            sheet.cell(row=row_number, column=column_number).value = column
            column_number += 1
        row_number += 1
        # Con este for se llenan las filas con los datoss
        for row in rows:
            column_number = 1
            for column in columns:
                sheet.cell(row=row_number, column=column_number).value = row[column_number - 1]
                column_number += 1
            row_number += 1
        # Se guarda el archivo de Excel en formato binario
        workbook.save(output)
        output.seek(0)

        return Response(output,
                        mimetype="application/ms-excel",
                        headers={"Content-Disposition": "attachment;filename=" + file_name})
