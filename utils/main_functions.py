# -*- coding: utf-8 -*-
"""
Created on Fri Sep 16 08:30:00 2022

@author: Jhonatan Martínez
"""

import inspect
import io
import os
from flask import flash, redirect, Response, url_for
from openpyxl import Workbook
from utils.data import messages


__this = 'main_functions.'


def create_directory(directory_name):
    """
    Crea un directorio en la ruta actual de la aplicación.

    Params:
        * **directory_name (Str):** Ruta o directorio a crear \n

    Returns:
        directory_name

    """
    try:
        # Validate if the directory don´t exists.
        if not os.path.exists(directory_name):
            os.makedirs(directory_name)
    except (OSError, Exception) as exc:
        # Variable error_message almacena la clase, el método y el error
        error_message = __this + inspect.stack()[0][3] + ': ' + str(exc)
        print(error_message)
    finally:
        return directory_name


def delete_file(path, fullpath=False):
    status = False
    try:
        if fullpath :
            os.remove(path)
        else:
            os.remove(get_current_path() + path)
        status = True
    except (OSError, Exception) as exc:
        # Variable error_message almacena la clase, el método y el error
        error_message = __this + inspect.stack()[0][3] + ': ' + str(exc)
        print(error_message)
    finally:
        return status


def download_excel(columns, rows, file_name):
    """
    Permite realizar la descarga de un archivo de Excel desde una consulta a una base de datos

    Params:
        * **columns (List):** Datos con los headers o nombres de las columnas. \n
        * **rows (List):** Datos a colocar en el archivo. \n
        * **file_name (Str):** Nombre con el que se va a descargar el archivo. \n

    Returns:
        Response
    """
    try:
        # Variable output para guardar el archivo en formato binario
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        row_number = 1
        column_number = 1
        # Con este for se llenan los headers o encabezados del archivo
        for column in columns:
            sheet.cell(row=row_number, column=column_number).value = column
            column_number += 1
        row_number += 1
        # Con este for se llenan las filas con los datos
        for row in rows:
            column_number = 1
            for column in columns:
                sheet.cell(row=row_number, column=column_number).value = row[column_number - 1]
                column_number += 1
            row_number += 1
        # Se guarda el archivo de Excel en formato binario
        workbook.save(output)
        output.seek(0)

        return Response(output,
                        mimetype="application/ms-excel",
                        headers={"Content-Disposition": "attachment;filename=" + file_name})
    except Exception as exc:
        # Variable error_message almacena la clase, el método y el error
        error_message = __this + inspect.stack()[0][3] + ': ' + str(exc)
        print(error_message)


def get_current_path():
    """
        Obtiene la ruta actual de la aplicación.

        Returns:
            **path (str):** Devuelve la ruta actual de la aplicación.

        """

    # Invoke class Answer.
    path = ""
    try:
        # Variable for the path.
        path = os.getcwd()
    except Exception as exc:
        # Variable error_message almacena la clase, el método y el error
        error_message = __this + inspect.stack()[0][3] + ': ' + str(exc)
        print(error_message)
    # Return answer object.
    return path


def organize_data(data_model, show_columns):
    """
    Organiza la data de modelos en listas con diccionarios para mostrar en los index

    Params:
            * **dataModel (Model):** Datos del modelo del ORM. \n
            * **show_columns (List):** Lista con los campos que se van a mostrar en el index. \n
            * **file_name (Str):** Nombre con el que se va a descargar el archivo. \n

        Returns:
            object_list
    """
    object_list = []
    try:
        # Recorrer cada fila devuelta por el modelo del ORM
        for row in data_model:
            # Variable dictionary almacena en formato diccionario cada fila del modelo del ORM
            dictionary = dict((column, getattr(row, column)) for column in row.__table__.columns.keys())
            delete_columns = []
            # Recorre las keys del diccionario
            for key in dictionary.keys():
                # Válida si la key está en las columnas que vamos a mostrar
                if key not in show_columns:
                    # Almacena las keys que no se van a mostrar
                    delete_columns.append(key)
            # For para eliminar las keys que no se van a mostrar
            for column in delete_columns:
                dictionary.pop(column)
            object_list.append(dictionary)
    except Exception as exc:
        # Variable error_message almacena la clase, el método y el error
        error_message = __this + inspect.stack()[0][3] + ': ' + str(exc)
        print(error_message)
    finally:
        return object_list


def show_message(id=0, message="", url="index"):
    flash(message=messages[id][0] if id != 0 else message,
          category=messages[id][1])
    return redirect(url_for(url))
